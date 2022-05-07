from os import listdir, makedirs
from os.path import isfile, join, isdir
from sys import exc_info
import argparse

from openpyxl import load_workbook


def parseArgs():
    parser = argparse.ArgumentParser()
    parser.add_argument("-d", "--data", required=True, help="directory containing data files directly from extracted Project .zip , the path to the project_Testing_the_clinica...folder", type=str)
    parser.add_argument("-o", "--output", required=True, help="output directory for results", type=str)
    #parser.add_argument("-b", "--bucket", required=False, help="S3 Bucket Name", type=str )

    return parser.parse_args()

def parseImmucorFile(csvDataLines=None, delimiter=',',csvFileName=None):
    currentCsvData = {}
    currentCsvData['manufacturer'] = 'immucor'
    currentCsvData['data']={}

    medianDatasetFound=False

    for csvDataLine in csvDataLines:
        #print('Line:' + str(csvDataLine))
        csvTokens = csvDataLine.split(delimiter)
        csvTokens = [token.strip().replace('"','') for token in csvTokens]

        if(medianDatasetFound):
            #print('Reading Median Dataset...')
            if(csvTokens[0] == 'Location'):
                # Header line
                currentCsvData['header_tokens']=csvTokens # TODO: should i trim off the first two here? Its location and sample. Do it later.
            elif(csvTokens[0] == ''):
                # End of the data set, we're done reading here.
                medianDatasetFound=False
                break # TODO: What if there are more than one? I assume there wont be
            else:
                # this row contains data
                # There is a comma in the location token. It's escaped by quotes but thats still annoying. I might need to combine the first and second tokens.
                if('(' in csvTokens[0] and ')' in csvTokens[1]):
                    csvTokens[0] = csvTokens[0] + '-' + csvTokens[1]
                    del csvTokens[1]

                sampleId=csvTokens[1]
                currentCsvData['data'][sampleId] = csvTokens
        else:
            if(csvTokens[0] == 'Samples'):
                #print('Sample Count Line:' + str(csvTokens))
                currentCsvData['expected_sample_count'] = int(csvTokens[1])
            elif(csvTokens[0] == 'DataType:' and csvTokens[1] == 'Median'):
                #print('Found the Median DataSet')
                medianDatasetFound = True
            else:
                pass# Do nothing with this line
    return currentCsvData

def determineManufacturer(csvDataLines=None, delimiter=',', csvFileName=None):
    for csvDataLine in csvDataLines:
        #print('Line:' + str(csvDataLine))
        csvTokens = csvDataLine.split(delimiter)
        csvTokens = [token.strip().replace('"','') for token in csvTokens]
        # Strange things with quotes and commas, watch out.
        #ProtocolDevelopingCompany,"Immucor Transplant Diag.,Inc"
        #"ProtocolDevelopingCompany","One Lambda, Inc."
        if len(csvTokens)>1:
            if(csvTokens[0]=='ProtocolDevelopingCompany' and 'Immucor' in csvTokens[1]):
                return 'immucor'
            elif(csvTokens[0]=='ProtocolDevelopingCompany' and 'One Lambda' in csvTokens[1]):
                return 'onelambda'

    # TODO: Handling edge cases, these are weird files:
    if(csvFileName in ['1668_1651227897977_ANTIBODY_CSV_ESPCRE_IM_casossamples20220408.csv'
        ,'1668_1651227898139_ANTIBODY_CSV_ESPCRE_IM_contolsamples20220408.csv'
        ,'1668_1651853325845_ANTIBODY_CSV_ESPCRE_IM_casossamples20220408.csv'
        ,'1668_1651853347374_ANTIBODY_CSV_ESPCRE_IM_contolsamples20220408.csv']):
        return 'immucor'
    elif(csvFileName in ['1668_1651853621467_ANTIBODY_CSV_ESPCRE_OL_grupo1_2_20220408.csv'
        ,'1668_1651853643342_ANTIBODY_CSV_ESPCRE_OL_grupo3_20220408.csv']):
        # These files look corrupt to me
        return None
    elif(csvFileName in []):
        return 'onelambda'

    print('could not determine manufacturer!')
    return None

def parseOneLambdaFile(csvDataLines=None, csvFileName=None, delimiter=','):
    currentCsvData = {}
    currentCsvData['manufacturer'] = 'onelambda'
    currentCsvData['data']={}

    trimmedMeanDatasetFound=False

    for csvDataLine in csvDataLines:
        #print('Line:' + str(csvDataLine))
        csvTokens = csvDataLine.split(delimiter)
        csvTokens = [token.strip().replace('"','') for token in csvTokens]

        if(trimmedMeanDatasetFound):
            #print('Reading Median Dataset...')
            if(csvTokens[0] == 'Location'):
                # Header line
                currentCsvData['header_tokens']=csvTokens # TODO: should i trim off the first two here? Its location and sample. Do it later.
            elif(csvTokens[0] == ''):
                # End of the data set, we're done reading here.
                trimmedMeanDatasetFound=False
                break # TODO: What if there are more than one? I assume there wont be
            else:
                # this row contains data
                # There is a comma in the location token. It's escaped by quotes but thats still annoying. I might need to combine the first and second tokens.
                if('(' in csvTokens[0] and ')' in csvTokens[1]):
                    csvTokens[0] = csvTokens[0] + '-' + csvTokens[1]
                    del csvTokens[1]

                sampleId=csvTokens[1]
                currentCsvData['data'][sampleId] = csvTokens
        else:
            if(csvTokens[0] == 'Samples'):
                #print('Sample Count Line:' + str(csvTokens))
                currentCsvData['expected_sample_count'] = int(csvTokens[1])
            elif(csvTokens[0] == 'DataType:' and csvTokens[1] == 'Trimmed Mean'):
                #print('Found the Trimmed Mean DataSet')
                trimmedMeanDatasetFound = True
            else:
                pass# Do nothing with this line
    return currentCsvData

def writeUnanalyzedFiles(unanalyzedFiles=None, outputDirectory=None, newline='\n', delimiter=','):
    outputFileName = join(outputDirectory,'UnanalyzedFiles.csv')
    print('writing un analyzed files:' + str(outputFileName))
    with open(outputFileName,'w') as outputFile:
        outputFile.write(delimiter.join(['filename','reason']) + newline)
        for unanalyzedFilename in unanalyzedFiles.keys():
            outputFile.write(delimiter.join([unanalyzedFilename,unanalyzedFiles[unanalyzedFilename]]) + newline)

def readCsvFiles(dataDirectory=None, outputDirectory=None):
    csvDirectory = join(dataDirectory,'ANTIBODY_CSV')
    print('Reading CSV files:' + str(csvDirectory))
    csvFileNames = sorted([f for f in listdir(csvDirectory) if isfile(join(csvDirectory, f)) and f.endswith('.csv')])
    print ('I found ' + str(len(csvFileNames)) + ' csv files.')

    csvData={}
    unanalyzedFiles = {}

    for csvFileName in csvFileNames:
    # for csvFileName in ['1659_1651225299574_ANTIBODY_CSV_DEUVER_220404-NHLA_20220404_142808.csv']:  # Immucor File
    # for csvFileName in ['1659_1651225315252_ANTIBODY_CSV_DEUVER_220413-NonHLA2_Auto1_20220413_114040.csv']: # One Lambda File

        # These files have a weird delimiter. I think they re-saved the file and need to re-export it. The files are full of extra dots.
        '''
        if(csvFileName in ['1676_1651237757189_ANTIBODY_CSV_NLDHEI_OL_group1_group2_post-Txsamples.csv'
            ,'1676_1651237951070_ANTIBODY_CSV_NLDHEI_OL_group1_group2_pre-Txsamples.csv'
            ,'1676_1651238170785_ANTIBODY_CSV_NLDHEI_OL_group3_post-Txsamples.csv'
            ,'1676_1651238309188_ANTIBODY_CSV_NLDHEI_OL_group3_pre-Txsamples.csv']):
            delimiter=';'
        else:
            delimiter=','
        '''
        delimiter = ','


        print('Reading file ' + str(csvFileName))
        currentCsvData = None
        try:
            with open(join(csvDirectory,csvFileName), 'r') as csvFile:
                csvDataLines=csvFile.readlines()
        except UnicodeDecodeError as e:
            #print('WARNING I COULD NOT READ THIS FILE BECAUSE OF UTF-8 ENCODING:' + str(csvFileName))
            print('Trying "ISO-8859-1" encoding')
            with open(join(csvDirectory, csvFileName), 'r', encoding='ISO-8859-1') as csvFile:
                csvDataLines = csvFile.readlines()
        #print('This file has ' + str(len(csvDataLines)) + ' lines.')

        manufacturer = determineManufacturer(csvDataLines=csvDataLines, csvFileName=csvFileName, delimiter=delimiter)
        if(manufacturer=='immucor'):
            currentCsvData = parseImmucorFile(csvDataLines=csvDataLines, csvFileName=csvFileName, delimiter=delimiter)
        elif(manufacturer=='onelambda'):
            currentCsvData = parseOneLambdaFile(csvDataLines=csvDataLines, csvFileName=csvFileName, delimiter=delimiter)
        else:
            #raise Exception('cannot determine file manufacturer:' + str(csvFileName))
            unanalyzedFiles[csvFileName] = 'cannot determine file manufacturer'
        if(currentCsvData is not None):
            csvData[csvFileName]={}
            csvData[csvFileName]['raw'] = currentCsvData

    writeUnanalyzedFiles(unanalyzedFiles=unanalyzedFiles, outputDirectory=outputDirectory)
    return csvData

def writeCombinedData(csvData=None, outputDirectory=None, delimiter =',', newline ='\n', rawOrClean=None):

    if(rawOrClean=='raw'):
        outputFileName = join(outputDirectory, 'CombinedRawData.csv')
    elif(rawOrClean=='clean'):
        outputFileName = join(outputDirectory, 'CombinedRawCleaned.csv')
    else:
        raise Exception('Am I printing raw or cleaned data?')

    print('writing combined raw data:' + str(outputFileName))
    with open(outputFileName, 'w') as outputFile:

        # immucor files, then one lambda files
        for csvFilename in sorted(csvData.keys()):
            if(csvData[csvFilename]['raw']['manufacturer'] == 'immucor'):
                #print('writing raw data for immucor file ' + str(csvFilename))
                outputFile.write(csvFilename + newline)
                outputFile.write('manufacturer:Immucor' + newline)
                outputFile.write(delimiter.join(csvData[csvFilename][rawOrClean]['header_tokens']) + newline)
                for sampleId in csvData[csvFilename][rawOrClean]['data'].keys():
                    outputFile.write(str(delimiter.join(csvData[csvFilename][rawOrClean]['data'][sampleId])) + newline)

                outputFile.write(newline)
            elif(csvData[csvFilename]['raw']['manufacturer'] == 'onelambda'):
                pass
            else:
                raise Exception('this file has no manufacturer:' + str(csvFilename))

        for csvFilename in sorted(csvData.keys()):
            if(csvData[csvFilename]['raw']['manufacturer'] == 'onelambda'):
                #print('writing raw data for onelambda file ' + str(csvFilename))
                outputFile.write(csvFilename + newline)
                outputFile.write('manufacturer:One Lambda' + newline)
                outputFile.write(delimiter.join(csvData[csvFilename][rawOrClean]['header_tokens']) + newline)
                for sampleId in csvData[csvFilename][rawOrClean]['data'].keys():
                    outputFile.write(str(delimiter.join(csvData[csvFilename][rawOrClean]['data'][sampleId])) + newline)
                outputFile.write(newline)

def cleanHeaders(rawHeaderTokens=None, manufacturer=None):
    cleanTokens = rawHeaderTokens.copy()

    negativeControlIndex = None
    positiveControlIndex = None

    if manufacturer == 'immucor':
        for headerTokenIndex, headerToken in enumerate(rawHeaderTokens):
            if (str(headerToken) == 'Location'):
                pass
            elif (str(headerToken) == 'Sample'):
                pass
            elif (str(headerToken) == 'Total Events'):
                pass
            elif (str(headerToken) == 'NC'):
                negativeControlIndex = headerTokenIndex
            elif (str(headerToken) == 'PC'):
                positiveControlIndex = headerTokenIndex
            elif (str(headerToken) == ''):
                pass
            elif (str(headerToken) == 'Notes'):
                pass
            elif (int(headerToken) == 501):
                cleanTokens[headerTokenIndex] = 'DEXI'
            elif (int(headerToken) == 502):
                cleanTokens[headerTokenIndex] = 'SNRPN'
            elif (int(headerToken) == 503):
                cleanTokens[headerTokenIndex] = 'LGALS3'
            elif (int(headerToken) == 504):
                cleanTokens[headerTokenIndex] = 'EMCN'
            elif (int(headerToken) == 505):
                cleanTokens[headerTokenIndex] = 'SSB'
            elif (int(headerToken) == 506):
                cleanTokens[headerTokenIndex] = 'IL21'
            elif (int(headerToken) == 507):
                cleanTokens[headerTokenIndex] = 'IL8'
            elif (int(headerToken) == 509):
                cleanTokens[headerTokenIndex] = 'VIM'
            elif (int(headerToken) == 510):
                cleanTokens[headerTokenIndex] = 'TUBA1B'
            elif (int(headerToken) == 511):
                cleanTokens[headerTokenIndex] = 'KRT8'
            elif (int(headerToken) == 512):
                cleanTokens[headerTokenIndex] = 'KRT18'
            elif (int(headerToken) == 513):
                cleanTokens[headerTokenIndex] = 'TUBULIN'
            elif (int(headerToken) == 514):
                cleanTokens[headerTokenIndex] = 'MYOSIN'
            elif (int(headerToken) == 515):
                cleanTokens[headerTokenIndex] = 'FLRT2'
            elif (int(headerToken) == 516):
                cleanTokens[headerTokenIndex] = 'COLLAGEN I'
            elif (int(headerToken) == 517):
                cleanTokens[headerTokenIndex] = 'COLLAGEN II'
            elif (int(headerToken) == 518):
                cleanTokens[headerTokenIndex] = 'COLLAGEN III'
            elif (int(headerToken) == 519):
                cleanTokens[headerTokenIndex] = 'COLLAGEN IV'
            elif (int(headerToken) == 520):
                cleanTokens[headerTokenIndex] = 'COLLAGEN V'
            elif (int(headerToken) == 521):
                cleanTokens[headerTokenIndex] = 'CSF2'
            elif (int(headerToken) == 522):
                cleanTokens[headerTokenIndex] = 'LGALS8'
            elif (int(headerToken) == 523):
                cleanTokens[headerTokenIndex] = 'SHC3'
            elif (int(headerToken) == 524):
                cleanTokens[headerTokenIndex] = 'STAT6'
            elif (int(headerToken) == 525):
                cleanTokens[headerTokenIndex] = 'AGRN'
            elif (int(headerToken) == 526):
                cleanTokens[headerTokenIndex] = 'PECR'
            elif (int(headerToken) == 527):
                cleanTokens[headerTokenIndex] = 'ATP5B'
            elif (int(headerToken) == 528):
                cleanTokens[headerTokenIndex] = 'APOL2'
            elif (int(headerToken) == 529):
                cleanTokens[headerTokenIndex] = 'CD40'
            elif (int(headerToken) == 530):
                cleanTokens[headerTokenIndex] = 'SNRPB2'
            elif (int(headerToken) == 531):
                cleanTokens[headerTokenIndex] = 'CXCL11'
            elif (int(headerToken) == 532):
                cleanTokens[headerTokenIndex] = 'GDNF'
            elif (int(headerToken) == 533):
                cleanTokens[headerTokenIndex] = 'IFNG'
            elif (int(headerToken) == 534):
                cleanTokens[headerTokenIndex] = 'CXCL9'
            elif (int(headerToken) == 535):
                cleanTokens[headerTokenIndex] = 'GSTT1'
            elif (int(headerToken) == 536):
                cleanTokens[headerTokenIndex] = 'ENO1'
            elif (int(headerToken) == 537):
                cleanTokens[headerTokenIndex] = 'GAPDH'
            elif (int(headerToken) == 538):
                cleanTokens[headerTokenIndex] = 'LMNA'
            elif (int(headerToken) == 539):
                cleanTokens[headerTokenIndex] = 'ACTIN'
            elif (int(headerToken) == 540):
                cleanTokens[headerTokenIndex] = 'NCL'
            elif (int(headerToken) == 541):
                cleanTokens[headerTokenIndex] = 'PRKCH'
            elif (int(headerToken) == 542):
                cleanTokens[headerTokenIndex] = 'PRKCZ'
            elif (int(headerToken) == 543):
                cleanTokens[headerTokenIndex] = 'PLA2R1'
            elif (int(headerToken) == 544):
                cleanTokens[headerTokenIndex] = 'FIBRONECTIN1'
            elif (int(headerToken) == 545):
                cleanTokens[headerTokenIndex] = 'HSPB1'
            elif (int(headerToken) == 546):
                cleanTokens[headerTokenIndex] = 'ROR1'
            elif (int(headerToken) == 547):
                cleanTokens[headerTokenIndex] = 'HARS'
            elif (int(headerToken) == 548):
                cleanTokens[headerTokenIndex] = 'ICAM1'
            elif (int(headerToken) == 549):
                cleanTokens[headerTokenIndex] = 'CGB5'
            elif (int(headerToken) == 550):
                cleanTokens[headerTokenIndex] = 'FAS'
            elif (int(headerToken) == 551):
                cleanTokens[headerTokenIndex] = 'P2RY11'
            elif (int(headerToken) == 552):
                cleanTokens[headerTokenIndex] = 'VEGFA'
            elif (int(headerToken) == 553):
                cleanTokens[headerTokenIndex] = 'VCL'
            elif (int(headerToken) == 554):
                cleanTokens[headerTokenIndex] = 'ARHGDIB'
            elif (int(headerToken) == 555):
                cleanTokens[headerTokenIndex] = 'PTPRO'
            elif (int(headerToken) == 556):
                cleanTokens[headerTokenIndex] = 'LPHN1'
            elif (int(headerToken) == 557):
                cleanTokens[headerTokenIndex] = 'COLLAGEN VI'
            elif (int(headerToken) == 558):
                cleanTokens[headerTokenIndex] = 'Thyroglobulin'
            elif (int(headerToken) == 559):
                cleanTokens[headerTokenIndex] = 'CCP'
            elif (int(headerToken) == 560):
                cleanTokens[headerTokenIndex] = 'Human Transferrin'
            elif (int(headerToken) == 561):
                cleanTokens[headerTokenIndex] = 'TUBB'

    elif manufacturer == 'onelambda':
        for headerTokenIndex, headerToken in enumerate(rawHeaderTokens):
            if (str(headerToken) == 'Location'):
                pass
            elif (str(headerToken) == 'Sample'):
                pass
            elif (str(headerToken) == 'Total Events'):
                pass
            elif (str(headerToken) == ''):
                pass
            elif (int(headerToken) == 1):
                cleanTokens[headerTokenIndex] = 'NC'
                negativeControlIndex = headerTokenIndex
            elif (int(headerToken) == 2):
                cleanTokens[headerTokenIndex] = 'PC'
                positiveControlIndex = headerTokenIndex
            elif (int(headerToken) == 3):
                cleanTokens[headerTokenIndex] = 'ENO1'
            elif (int(headerToken) == 4):
                cleanTokens[headerTokenIndex] = 'FLRT2'
            elif (int(headerToken) == 5):
                cleanTokens[headerTokenIndex] = 'VM'
            elif (int(headerToken) == 6):
                cleanTokens[headerTokenIndex] = 'TUBA1B'
            elif (int(headerToken) == 9):
                cleanTokens[headerTokenIndex] = 'IFIH1'
            elif (int(headerToken) == 12):
                cleanTokens[headerTokenIndex] = 'PTPRN'
            elif (int(headerToken) == 13):
                cleanTokens[headerTokenIndex] = 'AURKA'
            elif (int(headerToken) == 13):
                cleanTokens[headerTokenIndex] = 'AURKA'
            elif (int(headerToken) == 15):
                cleanTokens[headerTokenIndex] = 'PPIA'
            elif (int(headerToken) == 16):
                cleanTokens[headerTokenIndex] = 'EIF2A'
            elif (int(headerToken) == 21):
                cleanTokens[headerTokenIndex] = 'PECR'
            elif (int(headerToken) == 22):
                cleanTokens[headerTokenIndex] = 'PRKCH'
            elif (int(headerToken) == 24):
                cleanTokens[headerTokenIndex] = 'CXCL11'
            elif (int(headerToken) == 25):
                cleanTokens[headerTokenIndex] = 'CXCL10'
            elif (int(headerToken) == 27):
                cleanTokens[headerTokenIndex] = 'AGRIN'
            elif (int(headerToken) == 34):
                cleanTokens[headerTokenIndex] = 'GAPDH'
            elif (int(headerToken) == 37):
                cleanTokens[headerTokenIndex] = 'MYOSIN'
            elif (int(headerToken) == 38):
                cleanTokens[headerTokenIndex] = 'AGT'
            elif (int(headerToken) == 39):
                cleanTokens[headerTokenIndex] = 'CHAF1B'
            elif (int(headerToken) == 40):
                cleanTokens[headerTokenIndex] = 'PLA2R'
            elif (int(headerToken) == 47):
                cleanTokens[headerTokenIndex] = 'GSTT1'
            elif (int(headerToken) == 48):
                cleanTokens[headerTokenIndex] = 'LMNA'
            elif (int(headerToken) == 49):
                cleanTokens[headerTokenIndex] = 'PRKCZ'
            elif (int(headerToken) == 53):
                cleanTokens[headerTokenIndex] = 'LMNB1'
            elif (int(headerToken) == 61):
                cleanTokens[headerTokenIndex] = 'ARHGDIB'
            elif (int(headerToken) == 63):
                cleanTokens[headerTokenIndex] = 'HNRNPK'
            elif (int(headerToken) == 65):
                cleanTokens[headerTokenIndex] = 'IFNG'
            elif (int(headerToken) == 66):
                cleanTokens[headerTokenIndex] = 'REG3A'
            elif (int(headerToken) == 69):
                cleanTokens[headerTokenIndex] = 'NCL'
            elif (int(headerToken) == 73):
                cleanTokens[headerTokenIndex] = 'CD36'
            elif (int(headerToken) == 76):
                cleanTokens[headerTokenIndex] = 'TNFA'
            elif (int(headerToken) == 78):
                cleanTokens[headerTokenIndex] = 'CXCL9'
            elif (int(headerToken) == 81):
                cleanTokens[headerTokenIndex] = 'GDNF'
            elif (int(headerToken) == 44):
                cleanTokens[headerTokenIndex] = 'LG3'
            elif (int(headerToken) == 48):
                cleanTokens[headerTokenIndex] = 'LMNA'
            elif (int(headerToken) == 49):
                cleanTokens[headerTokenIndex] = 'PRKCZ'
            elif (int(headerToken) == 53):
                cleanTokens[headerTokenIndex] = 'LMNB1'
            elif (int(headerToken) == 61):
                cleanTokens[headerTokenIndex] = 'ARHGDIB'
            elif (int(headerToken) == 63):
                cleanTokens[headerTokenIndex] = 'HNRNPK'
            elif (int(headerToken) == 65):
                cleanTokens[headerTokenIndex] = 'IFNG'
            elif (int(headerToken) == 66):
                cleanTokens[headerTokenIndex] = 'REG3A'
            elif (int(headerToken) == 69):
                cleanTokens[headerTokenIndex] = 'NCL'
            elif (int(headerToken) == 73):
                cleanTokens[headerTokenIndex] = 'CD36'
            elif (int(headerToken) == 76):
                cleanTokens[headerTokenIndex] = 'TNFA'
            elif (int(headerToken) == 78):
                cleanTokens[headerTokenIndex] = 'CXCL9'
            elif (int(headerToken) == 81):
                cleanTokens[headerTokenIndex] = 'GDNF'
            elif (int(headerToken) == 44):
                cleanTokens[headerTokenIndex] = 'LG3'
            elif (int(headerToken) == 40):  # TODO: there are more beads, but some of these bead id repeat. There is a #40 in the twice?
                cleanTokens[headerTokenIndex] = 'COLLAGEN I'
            elif (int(headerToken) == 41):
                cleanTokens[headerTokenIndex] = 'COLLAGEN II'
            elif (int(headerToken) == 42):
                cleanTokens[headerTokenIndex] = 'COLLAGEN III'
            elif (int(headerToken) == 43):
                cleanTokens[headerTokenIndex] = 'COLLAGEN IV'
            elif (int(headerToken) == 44):
                cleanTokens[headerTokenIndex] = 'COLLAGEN V'
            elif (int(headerToken) == 45):
                cleanTokens[headerTokenIndex] = 'Fibronectin'
    else:
        raise Exception('This sample does not have a manufacturer')

    return cleanTokens, negativeControlIndex, positiveControlIndex

def subtractNegativeControls(negativeControlIndex=None, positiveControlIndex=None, rawDataRows=None, csvFileName=None):
    cleanedData={}
    for sampleId in rawDataRows.keys():
        rawDataRow = rawDataRows[sampleId]
        cleanDataRow = [None] * len(rawDataRow)
        for mfiIndex, mfi in enumerate(rawDataRow):
            # "Location" or "Sample", or control
            if(mfiIndex in [0,1, negativeControlIndex, positiveControlIndex]):
                cleanDataRow[mfiIndex]=rawDataRow[mfiIndex]
            elif (str(rawDataRow[mfiIndex]).upper() == 'NAN'):
                cleanDataRow[mfiIndex] = rawDataRow[mfiIndex]
            elif(str(rawDataRow[mfiIndex]).strip()==''):
                cleanDataRow[mfiIndex] = rawDataRow[mfiIndex]
            else:
                #print('csvFileName:' + str(csvFileName))
                #print('sampleid:' + str(sampleId))
                #print('rawData:' + rawDataRow[mfiIndex])
                #print('negativeControl:' + str(rawDataRow[negativeControlIndex]))
                cleanDataRow[mfiIndex] = str(float(rawDataRow[mfiIndex]) - float(rawDataRow[negativeControlIndex]))

                if(float(cleanDataRow[mfiIndex]) < 0):
                    cleanDataRow[mfiIndex] = '0'
        cleanedData[sampleId] = cleanDataRow

    return cleanedData

def convertAntigensAndSubtractControls(csvData=None):
    print('Converting Antigen Names and Subtracting Control Values')

    for csvFileName in csvData.keys():
        rawData = csvData[csvFileName]['raw']
        cleanedData = {}

        # Replace Headers with Antigen Names, and find the negative controls
        cleanedData['header_tokens'], cleanedData['negative_control_index'], cleanedData['positive_control_index'] = cleanHeaders(rawHeaderTokens=rawData['header_tokens'], manufacturer=rawData['manufacturer'])

        # Subtract the negative controls
        cleanedData['data'] = subtractNegativeControls(negativeControlIndex=cleanedData['negative_control_index'], positiveControlIndex=cleanedData['positive_control_index'], rawDataRows=rawData['data'], csvFileName=csvFileName)

        csvData[csvFileName]['clean'] = cleanedData

    return csvData

def readDataMatrices(dataDirectory=None, outputDirectory=None):
    dataMatrixDirectory = join(dataDirectory,'PROJECT_DATA_MATRIX')
    print('Reading Data Matrices:' + dataMatrixDirectory)
    excelFileNames = sorted([f for f in listdir(dataMatrixDirectory) if isfile(join(dataMatrixDirectory, f)) and f.endswith('.xlsx')])
    print('I found ' + str(len(excelFileNames)) + ' data matrices files.')

    excelData = {}
    unanalyzedMatrices = {}

    for excelFileName in excelFileNames:
        print('Reading file:' + str(excelFileName))
        xlWorkbook = load_workbook(filename=join(dataMatrixDirectory, excelFileName))
        dataSheet = xlWorkbook['Antibody_Data_Template']

        currentExcelData={}

        # Get the column indexes
        patientIdColumn=None
        rejectionColumn=None
        preTxSampleIdColumn=None
        preTxCsvImmucorColumn=None
        postTxSampleIdColumn=None
        postTxCsvImmucorColumn=None
        preTxCsvOneLambdaColumn=None
        postTxCsvOneLambdaColumn=None

        for colIndexRaw, col in enumerate(dataSheet.iter_cols()):
            if str(col[0].value).lower().strip()=='patient_identifier':
                patientIndexColumn = colIndexRaw
            elif str(col[0].value).lower().strip() == 'rejection':
                rejectionColumn = colIndexRaw
            elif str(col[0].value).lower().strip() == 'pre_tx_sample_id':
                preTxSampleIdColumn = colIndexRaw
            elif str(col[0].value).lower().strip() == 'pre_tx_csv_immucor':
                preTxCsvImmucorColumn = colIndexRaw
            elif str(col[0].value).lower().strip() == 'post_tx_sample_id':
                postTxSampleIdColumn = colIndexRaw
            elif str(col[0].value).lower().strip() == 'post_tx_csv_immucor':
                postTxCsvImmucorColumn = colIndexRaw
            elif str(col[0].value).lower().strip() == 'pre_tx_csv_onelambda':
                preTxCsvOneLambdaColumn = colIndexRaw
            elif str(col[0].value).lower().strip() == 'post_tx_csv_onelambda':
                postTxCsvOneLambdaColumn = colIndexRaw

        for rowIndexRaw, row in enumerate(dataSheet.iter_rows()):
            if(rowIndexRaw == 0):
                pass # header row
            else:
                patientId = row[patientIndexColumn].value
                if(patientId is not None and len(str(patientId).strip()) > 0):
                    currentExcelData[patientId] = {}
                    currentExcelData[patientId]['rejection']=row[rejectionColumn].value
                    currentExcelData[patientId]['pre_tx_sample_id']=row[preTxSampleIdColumn].value
                    currentExcelData[patientId]['pre_tx_csv_immucor']=row[preTxCsvImmucorColumn].value
                    currentExcelData[patientId]['post_tx_sample_id']=row[postTxSampleIdColumn].value
                    currentExcelData[patientId]['post_tx_csv_immucor']=row[postTxCsvImmucorColumn].value
                    currentExcelData[patientId]['pre_tx_csv_onelambda']=row[preTxCsvOneLambdaColumn].value
                    currentExcelData[patientId]['post_tx_csv_onelambda']=row[postTxCsvOneLambdaColumn].value

        excelData[excelFileName]=currentExcelData
    return excelData


def splitSamples(dataMatrixData=None, csvData=None):
    print('Splitting samples into reject and control')
    rejectionData = {}
    controlData = {}

    for excelFileName in dataMatrixData.keys():
        print('excel:' + str(excelFileName))
        for sampleId in dataMatrixData[excelFileName].keys():
            print('sampleId:' + str(sampleId))

            preTxData, postTxData = findSampleData(csvData=csvData, dataMatrixDataRow=dataMatrixData[excelFileName][sampleId])

    return rejectionData, controlData


def analyzeData(dataDirectory=None, outputDirectory=None):
    print('Looking for data files in ' + str(dataDirectory))

    # Loop Input .csv
    # Read Data with Sample IDs
    csvData=readCsvFiles(dataDirectory=dataDirectory, outputDirectory=outputDirectory)

    # write the raw data
    writeCombinedData(csvData=csvData, outputDirectory=outputDirectory, rawOrClean='raw')

    # Convert antigen names, using the lookup from Raymond
    csvData = convertAntigensAndSubtractControls(csvData=csvData)

    # write the modified/cleaned data
    writeCombinedData(csvData=csvData, outputDirectory=outputDirectory, rawOrClean='clean')

    # Read data matrix info
    dataMatrixData = readDataMatrices(dataDirectory=dataDirectory, outputDirectory=outputDirectory)

    # Split data rows into rejection and control samples
    rejectionData, controlData = splitSamples(dataMatrixData=dataMatrixData,csvData=csvData)

    #print('Writing files to ' + str(outputDirectory))


if __name__=='__main__':
    try:
        args=parseArgs()

        if not isdir(args.output):
            makedirs(args.output)

        analyzeData(dataDirectory=args.data, outputDirectory=args.output)

    except Exception:
        print ('Unexpected problem running tests:')
        print (str(exc_info()))
        raise